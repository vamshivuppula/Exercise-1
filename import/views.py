from django.shortcuts import render
import openpyxl
# Create your views here.
from openpyxl import load_workbook
from .models import PatientInformation

#def import_file(request):
#    if request.method == 'POST':
#        file = request.FILES['document']
##    return render(request, 'import.html')

def import_data(request):
    #wb = load_workbook(filename=request, data_only=True)
    path = '/home/administrator/Desktop/Exercise-1/exercise/source_data.xlsx'
    wb_obj = openpyxl.load_workbook(path)
    data_list = []
    data = test(wb_obj)
    data_list.extend(data)
    print(data,'data is')
    import_names(data_list)

def test(wb_obj):
    value_set = []
    for x in wb_obj:
        for row_tuple in x.iter_rows(
                min_row=2, max_col=x.max_column,
                values_only=True):
                value_set.append(row_tuple)
        #print(value_set,"This is in vjsvh")
        #print(value_set,'hiytrewq')

        return list(value_set)


def import_names(data_list):
    count = 1
    for names in data_list:
        #print(names[1],'final')
        name_data = PatientInformation(count,names[0],names[1],names[2],names[3])
        name_data.save()
        count = count +1
